# exporter.py
import os
import subprocess
from fpdf import FPDF

def exportar_txt(archivo_destino, contenido):
    """
    Exporta el contenido a un archivo TXT.
    
    Parámetros:
      archivo_destino (str): Ruta donde se guardará el archivo.
      contenido (str): Contenido a exportar.
    """
    with open(archivo_destino, "w", encoding="utf-8") as f:
        f.write(contenido)

def exportar_pdf(archivo_destino, contenido):
    """
    Exporta el contenido a un archivo PDF usando FPDF.
    
    Parámetros:
      archivo_destino (str): Ruta del archivo PDF a generar.
      contenido (str): Contenido a exportar.
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    # Dividimos el contenido en líneas y las agregamos al PDF
    for linea in contenido.splitlines():
        pdf.cell(0, 10, txt=linea, ln=True)
    pdf.output(archivo_destino)

def generar_mermaid(instrucciones):
    """
    Genera código en formato Mermaid.js para un diagrama de flujo.
    
    Parámetros:
      instrucciones (list): Lista de instrucciones o pasos.
      
    Retorna:
      str: Código Mermaid.js para representar un diagrama de flujo.
    """
    mermaid = "flowchart TD\n"
    for i, instr in enumerate(instrucciones, start=1):
        # Cada paso se define como un nodo
        mermaid += f"  step{i}[{instr}]\n"
        if i > 1:
            mermaid += f"  step{i-1} --> step{i}\n"
    return mermaid

def generar_documento_scrum(instrucciones, num_sprints):
    """
    Genera un documento SCRUM a partir de una lista de instrucciones y el número de sprints.
    Divide las instrucciones de forma equitativa entre los sprints.
    
    Parámetros:
      instrucciones (list): Lista de instrucciones extraídas.
      num_sprints (int): Número de sprints en los que se dividirán las instrucciones.
    
    Retorna:
      str: Documento SCRUM formateado.
    """
    if num_sprints < 1:
        num_sprints = 1
    total = len(instrucciones)
    if total == 0:
        return "No hay instrucciones para generar el documento SCRUM."
    
    # Calcular el tamaño base y el residuo para distribuir equitativamente
    sprint_size = total // num_sprints
    remainder = total % num_sprints

    sprints = {}
    start = 0
    for i in range(num_sprints):
        extra = 1 if i < remainder else 0
        end = start + sprint_size + extra
        sprints[f"Sprint {i+1}"] = instrucciones[start:end]
        start = end

    documento = ""
    for sprint, tareas in sprints.items():
        documento += f"{sprint}:\n"
        for tarea in tareas:
            documento += f"  - {tarea}\n"
        documento += "\n"
    return documento

def exportar_mermaid_diagrama(instrucciones, archivo_imagen):
    """
    Genera el diagrama Mermaid a partir de 'instrucciones' y lo exporta como imagen (PNG, JPG, SVG)
    usando mermaid-cli (mmdc).
    
    Parámetros:
      instrucciones (list): Lista de pasos o nodos del diagrama.
      archivo_imagen (str): Ruta donde se guardará la imagen (puede ser .png, .jpg o .svg).
    
    Requisitos:
      - Node.js instalado.
      - Mermaid CLI instalado globalmente (npm install -g @mermaid-js/mermaid-cli).
    """
    # 1. Generar el código Mermaid
    code = generar_mermaid(instrucciones)
    
    # 2. Crear un archivo temporal con extensión .mmd
    temp_mmd = "temp_mermaid.mmd"
    with open(temp_mmd, "w", encoding="utf-8") as f:
        f.write(code)
    
    # 3. Llamar a mmdc para convertir .mmd a la imagen
    comando = f'mmdc -i "{temp_mmd}" -o "{archivo_imagen}"'
    try:
        subprocess.run(comando, shell=True, check=True)
    except subprocess.CalledProcessError as e:
        raise RuntimeError("Error al generar la imagen Mermaid. Asegúrate de tener instalado Mermaid CLI.") from e
    finally:
        # 4. Eliminar el archivo temporal
        if os.path.exists(temp_mmd):
            os.remove(temp_mmd)

def generar_documento_calendario(instrucciones, num_sprints, total_days):
    """
    Genera un documento estilo calendario que distribuye las instrucciones entre sprints,
    asignando a cada sprint una cantidad de días basada en el total indicado.
    
    Parámetros:
      instrucciones (list): Lista de instrucciones extraídas.
      num_sprints (int): Número de sprints en los que se dividirán las instrucciones.
      total_days (int): Número total de días disponibles para realizar el trabajo.
    
    Retorna:
      str: Documento formateado en estilo calendario con sprints y días asignados.
    """
    if num_sprints < 1:
        num_sprints = 1
    total = len(instrucciones)
    if total == 0:
        return "No hay instrucciones para generar el documento calendario."
    
    # Distribuir las instrucciones entre los sprints (similar a generar_documento_scrum)
    sprint_size = total // num_sprints
    remainder = total % num_sprints

    sprints = {}
    start = 0
    for i in range(num_sprints):
        extra = 1 if i < remainder else 0
        end = start + sprint_size + extra
        sprints[f"Sprint {i+1}"] = instrucciones[start:end]
        start = end

    # Distribuir los días entre los sprints
    days_per_sprint = total_days // num_sprints
    days_remainder = total_days % num_sprints

    calendario = ""
    for i, (sprint, tasks) in enumerate(sprints.items()):
        extra_day = 1 if i < days_remainder else 0
        sprint_days = days_per_sprint + extra_day
        calendario += f"{sprint} - Duración: {sprint_days} día(s)\n"
        for task in tasks:
            calendario += f"  - {task}\n"
        calendario += "\n"
    return calendario

def exportar_calendario_excel(instrucciones, num_sprints, total_days, archivo_destino):
    """
    Genera un documento de calendario en formato Excel (.xlsx) que distribuye las instrucciones
    entre sprints, asignando a cada sprint una cantidad de días basada en el total indicado.
    Incluye una columna "Completado" con validación (Sí, No), una tabla resumen con fórmulas
    y un gráfico para visualizar el progreso.
    
    Parámetros:
      instrucciones (list): Lista de instrucciones extraídas.
      num_sprints (int): Número de sprints.
      total_days (int): Número total de días disponibles para el trabajo.
      archivo_destino (str): Ruta del archivo Excel (.xlsx) a generar.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Calendario"

    # Definir estilos
    header_fill = PatternFill(start_color="E5E5EA", end_color="E5E5EA", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    
    # Bordes finos para las celdas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados con la columna adicional "Completado"
    headers = ["Sprint", "Duración (días)", "Tarea", "Completado"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Verificar si hay instrucciones
    if num_sprints < 1:
        num_sprints = 1
    total = len(instrucciones)
    if total == 0:
        ws.append(["N/A", "N/A", "No hay instrucciones.", ""])
        wb.save(archivo_destino)
        return

    # Distribuir las instrucciones entre los sprints
    sprint_size = total // num_sprints
    remainder = total % num_sprints
    sprints = {}
    start = 0
    for i in range(num_sprints):
        extra = 1 if i < remainder else 0
        end = start + sprint_size + extra
        sprints[f"Sprint {i+1}"] = instrucciones[start:end]
        start = end

    # Distribuir los días entre los sprints
    days_per_sprint = total_days // num_sprints
    days_remainder = total_days % num_sprints

    # Escribir las filas para cada sprint y sus tareas con formato y filas alternadas
    current_row = 2
    fill_alternate = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for i, (sprint, tasks) in enumerate(sprints.items()):
        extra_day = 1 if i < days_remainder else 0
        sprint_days = days_per_sprint + extra_day
        if tasks:
            first = True
            for task in tasks:
                if first:
                    row_data = [sprint, sprint_days, task, ""]
                    first = False
                else:
                    row_data = ["", "", task, ""]
                ws.append(row_data)
                # Aplicar formato a la fila
                for col in range(1, 5):
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    # Alternar color de fondo en filas pares
                    if current_row % 2 == 0:
                        cell.fill = fill_alternate
                current_row += 1
        else:
            ws.append([sprint, sprint_days, "", ""])
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = cell_alignment
                cell.border = thin_border
                if current_row % 2 == 0:
                    cell.fill = fill_alternate
            current_row += 1

    # Agregar validación de datos a la columna "Completado" (columna D)
    dv = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    dv.error = 'Selecciona Sí o No'
    dv.errorTitle = 'Valor inválido'
    dv_range = f"D2:D{current_row-1}"
    ws.add_data_validation(dv)
    dv.add(dv_range)

    # Autoajuste de ancho de columnas según el contenido
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 4

    # --- Agregar tabla resumen ---
    summary_start = current_row + 2
    ws.cell(row=summary_start, column=1, value="Resumen de Tareas").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=2)

    # Encabezados del resumen
    ws.cell(row=summary_start+1, column=1, value="Métrica").font = header_font
    ws.cell(row=summary_start+1, column=2, value="Cantidad").font = header_font
    ws.cell(row=summary_start+1, column=1).fill = header_fill
    ws.cell(row=summary_start+1, column=2).fill = header_fill
    ws.cell(row=summary_start+1, column=1).alignment = header_alignment
    ws.cell(row=summary_start+1, column=2).alignment = header_alignment
    ws.cell(row=summary_start+1, column=1).border = thin_border
    ws.cell(row=summary_start+1, column=2).border = thin_border

    last_data_row = current_row - 1
    ws.cell(row=summary_start+2, column=1, value="Total Tareas")
    ws.cell(row=summary_start+2, column=2, 
            value=f"=COUNTA(C2:C{last_data_row})")
    ws.cell(row=summary_start+3, column=1, value="Completadas")
    ws.cell(row=summary_start+3, column=2, 
            value=f"=COUNTIF(D2:D{last_data_row}, \"Sí\")")
    ws.cell(row=summary_start+4, column=1, value="Pendientes")
    ws.cell(row=summary_start+4, column=2, 
            value=f"=COUNTIF(D2:D{last_data_row}, \"No\")")

    # Formato para las celdas del resumen
    for row in range(summary_start+2, summary_start+5):
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            cell.alignment = header_alignment
            cell.border = thin_border

    # --- Agregar gráfico de pastel ---
    pie = PieChart()
    pie.title = "Progreso de Tareas"
    labels = Reference(ws, min_col=1, min_row=summary_start+3, max_row=summary_start+4)
    data = Reference(ws, min_col=2, min_row=summary_start+3, max_row=summary_start+4)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 7    # Altura del gráfico
    pie.width = 7     # Ancho del gráfico

    # Ubicar el gráfico al lado del resumen
    ws.add_chart(pie, f"E{summary_start}")

    wb.save(archivo_destino)

